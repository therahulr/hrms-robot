import calendar
import datetime
import os
import random
import shelve
from datetime import date
from datetime import timedelta

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import *

from resources.test_data.scrape_data.data_dictionary import *


class TimeWiseAttendance:
    def __init__(self):
        self.month_dict = month_dict
        self.holiday_list = holiday_list

        self.today = date.today()
        self.current_year = date.today().year
        self.month = int(input("Enter month number : "))
        self.start_date = self.today.replace(day=1, month=self.month)
        self.last_day = calendar.monthrange(self.current_year, self.month)[1]
        self.last_date = self.today.replace(day=self.last_day, month=self.month)
        self.date_list = [d.strftime("%d-%m-%Y") for d in
                          (self.start_date + timedelta(days=x) for x in
                           range((self.last_date - self.start_date).days + 1))]

        print("Enter one or more weekend day's number separated by comma (For default press 0 or press ENTER)")
        self.user_input = [x for x in input().split(',')]

    def read_db(self):
        emp_db = os.getcwd() + "/scrape_data/employee_list_db/empdb"
        with shelve.open(filename=emp_db) as emp:
            data = emp['employees']
        return data

    def holidays_dates(self):
        holiday_d = []
        if holiday_list[self.month]:
            holiday_date = [i for i in holiday_list[self.month]]
            holiday_d = []
            for i in holiday_date:
                holiday_d.append(self.today.replace(day=i, month=self.month))
        return [i.strftime("%Y-%m-%d") for i in holiday_d]

    # Finding Weekends

    def get_no_of_weekends(self):

        weekends = []

        if len(self.user_input) == 1:
            for num in self.user_input:
                if num == str(0):

                    weekends = [7]
                elif num == "":
                    weekends = [7]
                elif str(1) <= num <= str(7):
                    weekends = self.user_input
                else:
                    print("Wrong data entered")

        elif 1 < len(self.user_input) <= 7:  # [3,5] If a user selects more than one weekend
            if any(str(1) > val > str(7) for val in self.user_input):
                print("You cannot enter a day number greater than 7")
            else:
                weekends = self.user_input
        else:
            print("Wrong data entered. Please try again")
        return weekends

    def total_weekends_days(self):
        today = datetime.date.today()
        start = today.replace(day=1, month=self.month)

        last_day = calendar.monthrange(self.current_year, self.month)[1]
        end = today.replace(day=last_day, month=self.month)

        weekends = []
        wknds = self.get_no_of_weekends()
        for d in wknds:
            weekend_date = [weekend_date.strftime("%d") for weekend_date in
                            [start + datetime.timedelta(day) for day in range(0, (end - start).days + 1)] if
                            weekend_date.isoweekday() == int(d)]
            [weekends.append(int(i)) for i in weekend_date]
        weekends.sort()
        #     return weekends
        weekend_dates = []

        for i in weekends:
            weekend_dates.append(today.replace(day=i, month=self.month))
        return [i.strftime("%d-%m-%Y") for i in weekend_dates]

    def get_total_off_days(self):
        total_off_days = list(set(self.total_weekends_days() + self.holidays_dates()))
        total_off_days.sort()
        return total_off_days

    def get_total_work_days(self):
        total_work_days = [day for day in self.date_list if day not in self.get_total_off_days()]
        return total_work_days

    """ Date and Time format"""
    frmt = '%H:%M:%S'

    intime_start = "09:00:00"
    intime_end = "09:30:00"

    out_time_start = "18:00:00"
    out_time_end = "18:45:00"

    half_time_start = "13:00:00"
    half_time_end = "14:30:00"

    def gen_random_time(self, start, end):
        s_time = datetime.datetime.strptime(start, self.frmt)
        e_time = datetime.datetime.strptime(end, self.frmt)
        time_diff = e_time - s_time
        return random.random() * time_diff + s_time

    def out_time_with_half_day(self):
        out_time_data = [self.gen_random_time(self.out_time_start, self.out_time_end).strftime(self.frmt),
                         self.gen_random_time(self.half_time_start, self.half_time_end).strftime(self.frmt)]
        return np.random.choice(out_time_data, p=[0.7, 0.3], size=1)

    def generate_excel_file(self):

        emp = self.read_db()
        total_work_days = self.get_total_work_days()
        # Excel File
        sheet_directory = "attendance_timewise/"
        wb = Workbook()
        ws = wb.create_sheet("data", 0)
        ws.column_dimensions['A'].width = 17
        ws.column_dimensions['B'].width = 23
        ws.column_dimensions['C'].width = 17
        ws.column_dimensions['D'].width = 17
        ws.column_dimensions['E'].width = 23
        start_row = 1
        start_col = 1
        filename = sheet_directory + month_dict[str(self.month)] + "_" + str(self.current_year) + ".xlsx"

        # Styles
        style1 = NamedStyle(name="style1")
        style1.font = Font(size=12, bold=True)
        style1.fill = PatternFill(fill_type="solid", fgColor="99CCFF")
        style1.alignment = Alignment(horizontal="center")
        style1.border = Border(left=Side(style="thin"),
                               right=Side(style="thin"),
                               top=Side(style="thin"),
                               bottom=Side(style="thin"))

        style2 = NamedStyle(name="style2")
        style2.alignment = Alignment(horizontal="center")
        style2.border = Border(left=Side(style="thin"),
                               right=Side(style="thin"),
                               top=Side(style="thin"),
                               bottom=Side(style="thin"))

        # Create Sheet and Fill Headers
        emp_id_cell = ws.cell(row=start_row, column=start_col)
        marked_on_cell = ws.cell(row=start_row, column=start_col + 1)
        intime_cell = ws.cell(row=start_row, column=start_col + 2)
        outtime_cell = ws.cell(row=start_row, column=start_col + 3)
        marked_by_cell = ws.cell(row=start_row, column=start_col + 4)

        emp_id_cell.style = style1
        marked_on_cell.style = style1
        intime_cell.style = style1
        outtime_cell.style = style1
        marked_by_cell.style = style1

        emp_id_cell.value = "employeeId"
        marked_on_cell.value = "attendanceMarkedOn"
        intime_cell.value = "inTime"
        outtime_cell.value = "outTime"
        marked_by_cell.value = "attendanceMarkedBy"

        # Fill Employee ID, Working dates and Marked by ID
        s_row = 2

        for dates in range(0, len(self.get_total_work_days())):

            for emp_id in range(0, len(emp)):
                ws.cell(row=s_row + emp_id, column=start_col).style = style2
                ws.cell(row=s_row + emp_id, column=start_col + 1).style = style2
                ws.cell(row=s_row + emp_id, column=start_col + 2).style = style2
                ws.cell(row=s_row + emp_id, column=start_col + 3).style = style2
                ws.cell(row=s_row + emp_id, column=start_col + 4).style = style2

                ws.cell(row=s_row + emp_id, column=start_col).value = str(emp[emp_id].get('id'))
                ws.cell(row=s_row + emp_id, column=start_col + 1).value = total_work_days[dates]
                ws.cell(row=s_row + emp_id, column=start_col + 2).value = self.gen_random_time(self.intime_start,
                                                                                               self.intime_end).strftime(
                    self.frmt)
                ws.cell(row=s_row + emp_id, column=start_col + 3).value = self.out_time_with_half_day()[0]
                ws.cell(row=s_row + emp_id, column=start_col + 4).value = "SSPL35119"

            s_row = s_row + len(emp)

        wb.save(filename)
        print("Excel file saved successfully.\nFile Path : {}".format(filename))


if __name__ == "__main__":
    obj = TimeWiseAttendance()
    obj.generate_excel_file()
