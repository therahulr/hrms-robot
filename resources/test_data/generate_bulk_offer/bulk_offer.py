import string
from datetime import datetime

from faker import Faker
from openpyxl import Workbook
from openpyxl.styles import *

from resources.test_data.generate_data.generate_test_data import *

fake = Faker('en_IN')


def random_string(length):
    """
    This function will generate random string of the specified length
    :param length:
    :return:
    """
    rand_string = ''.join(random.choice(string.ascii_uppercase) for i in range(length))
    return rand_string


def get_address():
    return fake.address()


def get_ems_data():
    """

    :rtype: object
    """
    first_name = fake.first_name()
    last_name = fake.last_name()
    middle_name = first_name[:3] + last_name.lower()[:3]
    dob = fake.date_between(start_date='-50y', end_date='-20y').strftime("%d-%m-%Y")
    official_email = first_name.lower() + last_name[:2].lower() + "@beyondhr.cloud"
    employee_id = "SSPL" + str(random.randint(1, 9)) + fake.numerify(text='##') + datetime.now().strftime("%S")
    gender = get_gender()
    category = get_category()
    marital_status = get_marital_status()
    pan_no = random_string(4) + last_name[:1] + fake.numerify(text='####') + random_string(1)
    aadhar_no = fake.numerify(text='9#####') + fake.numerify(text='######')
    phone = fake.numerify(text='98########')  # generates 10 digit mobile number (starts with 98)
    designation = get_designation()
    department = get_department()
    doj = fake.date_this_year(before_today=False, after_today=True).strftime("%d-%m-%Y")
    location = get_location()
    employee_type = get_employee_type()
    reportee_manager = get_reportee_manager()
    address_type = "Current"
    primary_address = get_address()
    landmark = get_address()
    street = get_address()
    country = get_country()
    state = "Bihar"
    city = "Madhubani"
    pin_code = str(random.randint(4, 9)) + fake.numerify(text='#####')
    fax = fake.numerify(text='98######')
    grade = get_employee_grade()
    employee_skill = get_employee_skill()

    data = {'fname': first_name,
            'mname': middle_name,
            'lname': last_name,
            'dob': dob,
            'email': official_email,
            'empid': employee_id,
            'gender': gender,
            'category': category,
            'marital_status': marital_status,
            'pan_no': pan_no,
            'aadhar': int(aadhar_no),
            'phone': int(phone),
            'designation': designation,
            'department': department,
            'doj': doj,
            'location': location,
            'employee_type': employee_type,
            'reportee_man': reportee_manager,
            'addr_type': address_type,
            'primary_address': primary_address,
            'landmark': landmark,
            'street': street,
            'country': country,
            'state': state,
            'city': city,
            'pin': int(pin_code),
            'fax': fax,
            'grade': grade,
            'employee_skill': employee_skill
            }
    return data


# Creating excel
def generate_bulk_offer():
    print("Enter number of candidates : ")
    num = int(input())

    sheet_directory = "../bulk_offer_data/"
    wb = Workbook()
    ws = wb.create_sheet("Worksheet", 0)
    start_row = 1
    start_col = 1
    filename = sheet_directory + str(num) + " Offer Bulk Data " + datetime.now().strftime(
        "%d_%b_%y") + ".xlsx"
    fname_cell = ws.cell(row=start_row, column=start_col)
    mname_cell = ws.cell(row=start_row, column=start_col + 1)
    lname_cell = ws.cell(row=start_row, column=start_col + 2)
    dob_cell = ws.cell(row=start_row, column=start_col + 3)
    email_cell = ws.cell(row=start_row, column=start_col + 4)
    empid_cell = ws.cell(row=start_row, column=start_col + 5)
    gender_cell = ws.cell(row=start_row, column=start_col + 6)
    category_cell = ws.cell(row=start_row, column=start_col + 7)
    marital_status_cell = ws.cell(row=start_row, column=start_col + 8)
    pan_cell = ws.cell(row=start_row, column=start_col + 9)
    aadhar_cell = ws.cell(row=start_row, column=start_col + 10)
    phone_cell = ws.cell(row=start_row, column=start_col + 11)
    designation_cell = ws.cell(row=start_row, column=start_col + 12)
    department_cell = ws.cell(row=start_row, column=start_col + 13)
    doj_cell = ws.cell(row=start_row, column=start_col + 14)
    location_cell = ws.cell(row=start_row, column=start_col + 15)
    emptype_cell = ws.cell(row=start_row, column=start_col + 16)
    reportee_man_cell = ws.cell(row=start_row, column=start_col + 17)
    address_type_cell = ws.cell(row=start_row, column=start_col + 18)
    primary_add_cell = ws.cell(row=start_row, column=start_col + 19)

    landmark_cell = ws.cell(row=start_row, column=start_col + 20)
    street_cell = ws.cell(row=start_row, column=start_col + 21)
    country_cell = ws.cell(row=start_row, column=start_col + 22)
    state_cell = ws.cell(row=start_row, column=start_col + 23)
    city_cell = ws.cell(row=start_row, column=start_col + 24)
    pin_cell = ws.cell(row=start_row, column=start_col + 25)
    fax_cell = ws.cell(row=start_row, column=start_col + 26)
    grade_cell = ws.cell(row=start_row, column=start_col + 27)
    skill_cell = ws.cell(row=start_row, column=start_col + 28)

    # Setting column width
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 28
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 25
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 25
    ws.column_dimensions['N'].width = 25
    ws.column_dimensions['O'].width = 15
    ws.column_dimensions['P'].width = 15
    ws.column_dimensions['Q'].width = 20
    ws.column_dimensions['R'].width = 20
    ws.column_dimensions['S'].width = 30
    ws.column_dimensions['T'].width = 30
    ws.column_dimensions['U'].width = 15
    ws.column_dimensions['V'].width = 15
    ws.column_dimensions['W'].width = 20
    ws.column_dimensions['X'].width = 15
    ws.column_dimensions['Y'].width = 15
    ws.column_dimensions['Z'].width = 20
    ws.column_dimensions['AA'].width = 20
    ws.column_dimensions['AB'].width = 20
    ws.column_dimensions['AC'].width = 20

    # Creating Styles to apply in excel
    header_style = NamedStyle("header_style")
    header_style.font = Font(size=13, bold=True)
    header_style.fill = PatternFill(fill_type="solid", fgColor="99CCFF")
    header_style.alignment = Alignment(horizontal="center")
    header_style.border = Border(left=Side(style="thin"),
                                 right=Side(style="thin"),
                                 top=Side(style="thin"),
                                 bottom=Side(style="thin"))

    data_style = NamedStyle("data_style")
    data_style.font = Font(size=11, bold=False)
    data_style.alignment = Alignment(horizontal="center")
    data_style.border = Border(left=Side(style="thin"),
                               right=Side(style="thin"),
                               top=Side(style="thin"),
                               bottom=Side(style="thin"))

    fname_cell.style = header_style
    fname_cell.value = "firstName"
    mname_cell.style = header_style
    mname_cell.value = "middleName"
    lname_cell.style = header_style
    lname_cell.value = "lastName"
    dob_cell.style = header_style
    dob_cell.value = "dob"
    email_cell.style = header_style
    email_cell.value = "officialEmail"
    empid_cell.style = header_style
    empid_cell.value = "employeeId"
    gender_cell.style = header_style
    gender_cell.value = "gender"
    category_cell.style = header_style
    category_cell.value = "category"
    marital_status_cell.style = header_style
    marital_status_cell.value = "maritalStatus"
    pan_cell.style = header_style
    pan_cell.value = "panNo"
    aadhar_cell.style = header_style
    aadhar_cell.value = "adhaarToken"
    phone_cell.style = header_style
    phone_cell.value = "phone"
    designation_cell.style = header_style
    designation_cell.value = "designation"
    department_cell.style = header_style
    department_cell.value = "department"
    doj_cell.style = header_style
    doj_cell.value = "doj"
    location_cell.style = header_style
    location_cell.value = "location"
    emptype_cell.style = header_style
    emptype_cell.value = "employeeType"
    reportee_man_cell.style = header_style
    reportee_man_cell.value = "reporteeManager"
    address_type_cell.style = header_style
    address_type_cell.value = "addressType"

    primary_add_cell.style = header_style
    primary_add_cell.value = "primaryAddress"
    landmark_cell.style = header_style
    landmark_cell.value = "landMark"
    street_cell.style = header_style
    street_cell.value = "street"
    country_cell.style = header_style
    country_cell.value = "country"
    state_cell.style = header_style
    state_cell.value = "state"
    city_cell.style = header_style
    city_cell.value = "city"
    pin_cell.style = header_style
    pin_cell.value = "pinCode"
    fax_cell.style = header_style
    fax_cell.value = "fax"
    grade_cell.style = header_style
    grade_cell.value = "grade"
    fax_cell.style = header_style
    skill_cell.style = header_style
    skill_cell.value = "employeeSkilled"

    # Writing data into excel file

    for i in range(1, num):
        data = get_ems_data()
        ws.cell(row=start_row + i, column=start_col).style = data_style
        ws.cell(row=start_row + i, column=start_col).value = data['fname']

        ws.cell(row=start_row + i, column=start_col + 1).style = data_style
        ws.cell(row=start_row + i, column=start_col + 1).value = data['mname']

        ws.cell(row=start_row + i, column=start_col + 2).style = data_style
        ws.cell(row=start_row + i, column=start_col + 2).value = data['lname']

        ws.cell(row=start_row + i, column=start_col + 3).style = data_style
        ws.cell(row=start_row + i, column=start_col + 3).value = data['dob']

        ws.cell(row=start_row + i, column=start_col + 4).style = data_style
        ws.cell(row=start_row + i, column=start_col + 4).value = data['email']

        ws.cell(row=start_row + i, column=start_col + 5).style = data_style
        ws.cell(row=start_row + i, column=start_col + 5).value = data['empid']

        ws.cell(row=start_row + i, column=start_col + 6).style = data_style
        ws.cell(row=start_row + i, column=start_col + 6).value = data['gender']

        ws.cell(row=start_row + i, column=start_col + 7).style = data_style
        ws.cell(row=start_row + i, column=start_col + 7).value = data['category']

        ws.cell(row=start_row + i, column=start_col + 8).style = data_style
        ws.cell(row=start_row + i, column=start_col + 8).value = data['marital_status']

        ws.cell(row=start_row + i, column=start_col + 9).style = data_style
        ws.cell(row=start_row + i, column=start_col + 9).value = data['pan_no']

        ws.cell(row=start_row + i, column=start_col + 10).style = data_style
        ws.cell(row=start_row + i, column=start_col + 10).value = data['aadhar']

        ws.cell(row=start_row + i, column=start_col + 11).style = data_style
        ws.cell(row=start_row + i, column=start_col + 11).value = data['phone']

        ws.cell(row=start_row + i, column=start_col + 12).style = data_style
        ws.cell(row=start_row + i, column=start_col + 12).value = data['designation']

        ws.cell(row=start_row + i, column=start_col + 13).style = data_style
        ws.cell(row=start_row + i, column=start_col + 13).value = data['department']

        ws.cell(row=start_row + i, column=start_col + 14).style = data_style
        ws.cell(row=start_row + i, column=start_col + 14).value = data['doj']

        ws.cell(row=start_row + i, column=start_col + 15).style = data_style
        ws.cell(row=start_row + i, column=start_col + 15).value = data['location']

        ws.cell(row=start_row + i, column=start_col + 16).style = data_style
        ws.cell(row=start_row + i, column=start_col + 16).value = data['employee_type']

        ws.cell(row=start_row + i, column=start_col + 17).style = data_style
        ws.cell(row=start_row + i, column=start_col + 17).value = data['reportee_man']

        ws.cell(row=start_row + i, column=start_col + 18).style = data_style
        ws.cell(row=start_row + i, column=start_col + 18).value = data['addr_type']

        ws.cell(row=start_row + i, column=start_col + 19).style = data_style
        ws.cell(row=start_row + i, column=start_col + 19).value = data['primary_address']

        ws.cell(row=start_row + i, column=start_col + 20).style = data_style
        ws.cell(row=start_row + i, column=start_col + 20).value = data['landmark']

        ws.cell(row=start_row + i, column=start_col + 21).style = data_style
        ws.cell(row=start_row + i, column=start_col + 21).value = data['street']

        ws.cell(row=start_row + i, column=start_col + 22).style = data_style
        ws.cell(row=start_row + i, column=start_col + 22).value = data['country']

        ws.cell(row=start_row + i, column=start_col + 23).style = data_style
        ws.cell(row=start_row + i, column=start_col + 23).value = data['state']

        ws.cell(row=start_row + i, column=start_col + 24).style = data_style
        ws.cell(row=start_row + i, column=start_col + 24).value = data['city']

        ws.cell(row=start_row + i, column=start_col + 25).style = data_style
        ws.cell(row=start_row + i, column=start_col + 25).value = data['pin']

        ws.cell(row=start_row + i, column=start_col + 26).style = data_style
        ws.cell(row=start_row + i, column=start_col + 26).value = data['fax']

        ws.cell(row=start_row + i, column=start_col + 27).style = data_style
        ws.cell(row=start_row + i, column=start_col + 27).value = data['grade']

        ws.cell(row=start_row + i, column=start_col + 28).style = data_style
        ws.cell(row=start_row + i, column=start_col + 28).value = data['employee_skill']

    wb.save(filename)


if __name__ == "__main__":
    generate_bulk_offer()
    print("Data generated success!")
