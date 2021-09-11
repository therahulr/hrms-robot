import os
import random
import shelve

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import *
from openpyxl.worksheet.dimensions import ColumnDimension

from scrapedata.datetime_util import Attendance
from test_data.scrape_data.data_dictionary import *


class GenerateAttendance(Attendance):

    def __init__(self):
        super().__init__()
        self.days = days
        self.month_dict = month_dict
        self.holiday_list = holiday_list
        self.sheet_directory = "daywise_attendance_sheet/"
        self.wb = Workbook()
        self.ws = self.wb.create_sheet("Attendance Sheet", 0)
        self.ws.column_dimensions['A'].width = 20
        self.ws.column_dimensions['B'].width = 20
        self.start_row = 1
        self.start_col = 1
        self.filename = self.sheet_directory + self.month_dict[str(self.month)] + "_" + str(self.current_year) + ".xlsx"

    def read_db(self):
        emp_db = os.getcwd() + "/scrape_data/employee_list_db/empdb"
        with shelve.open(filename=emp_db) as emp:
            data = emp['employees']
        return data

    def create_excel_file(self):

        weekends_dates = self.get_no_of_weekends()
        print("Selected weekend days are: ", [self.days[int(x)] + " " for x in weekends_dates])
        year = self.current_year
        weekends = self.total_weekends_days()
        ems_data = self.read_db()

        # Fill Excel Headers
        ColumnDimension(worksheet=self.ws, auto_size=True, index='A')
        ColumnDimension(worksheet=self.ws, auto_size=True, index='B')
        year_cell = self.ws.cell(row=self.start_row, column=self.start_col)
        month_cell = self.ws.cell(row=self.start_row + 1, column=self.start_col)
        empid_cell = self.ws.cell(row=self.start_row + 2, column=self.start_col)
        year_val_cell = self.ws.cell(row=self.start_row, column=self.start_col + 1)
        month_val_cell = self.ws.cell(row=self.start_row + 1, column=self.start_col + 1)
        name_cell = self.ws.cell(row=self.start_row + 2, column=self.start_col + 1)

        # Styling cells
        style1 = NamedStyle(name="style1")
        style1.font = Font(size=12, bold=True)
        style1.fill = PatternFill(fill_type="solid", fgColor="99CCFF")
        style1.alignment = Alignment(horizontal="center")
        style1.border = Border(left=Side(style="thin"),
                               right=Side(style="thin"),
                               top=Side(style="thin"),
                               bottom=Side(style="thin"))

        style2 = NamedStyle(name="style2")
        style2.font = Font(bold=True)
        style2.fill = PatternFill(fill_type="solid", fgColor="99CCFF")
        style2.alignment = Alignment(horizontal="center")
        style2.border = Border(left=Side(style="thin"),
                               right=Side(style="thin"),
                               top=Side(style="thin"),
                               bottom=Side(style="thin"))

        style3 = NamedStyle(name="style3")
        style3.alignment = Alignment(horizontal="center")
        style3.border = Border(left=Side(style="thin"),
                               right=Side(style="thin"),
                               top=Side(style="thin"),
                               bottom=Side(style="thin"))

        style4 = NamedStyle(name="style4")
        style4.font = Font(bold=True)
        style4.fill = PatternFill(fill_type="solid", fgColor="99CCFF")
        style4.alignment = Alignment(horizontal="center")
        style4.border = Border(left=Side(style="thin"),
                               right=Side(style="thin"),
                               top=Side(style="thin"),
                               bottom=Side(style="thin"))

        # Filling Data
        year_cell.style = style1
        month_cell.style = style1
        empid_cell.style = style1
        year_cell.value = "Year"
        month_cell.value = "Month"
        empid_cell.value = "EmployeeID"
        name_cell.style = style1
        name_cell.value = "Name (optional)"
        year_val_cell.style = style4
        year_val_cell.value = year
        month_val_cell.style = style4
        month_val_cell.value = self.month

        # Fill all date header
        for dates_header in range(1, self.get_no_of_days_in_month() + 1):
            self.ws.cell(row=self.start_row + 2, column=self.start_col + 1 + dates_header).value = dates_header
            self.ws.cell(row=self.start_row + 2, column=self.start_col + 1 + dates_header).style = style2

        # Fill Employee ID and Name
        for name in range(0, len(ems_data)):
            self.ws.cell(row=self.start_row + 3 + name, column=self.start_col).style = style3
            self.ws.cell(row=self.start_row + 3 + name, column=self.start_col).value = str(ems_data[name].get('id'))
            self.ws.cell(row=self.start_row + 3 + name, column=self.start_col + 1).style = style3
            self.ws.cell(row=self.start_row + 3 + name, column=self.start_col + 1).value = ems_data[name].get('name')

        # Fill Holiday
        for hol in range(0, len(self.get_holidays())):
            hol_col = self.get_holidays()[hol] + 2
            for hol_abbr in range(0, len(ems_data)):
                self.ws.cell(row=self.start_row + 3 + hol_abbr, column=hol_col).style = style3
                self.ws.cell(row=self.start_row + 3 + hol_abbr, column=hol_col).fill = PatternFill(fill_type="solid",
                                                                                                   fgColor="FF9999")
                self.ws.cell(row=self.start_row + 3 + hol_abbr, column=hol_col).value = "HOL"

        # Fill Weekend Abbr
        for wknds in range(0, len(weekends)):
            col_val = weekends[wknds] + 2
            for weekends_abbr in range(0, len(ems_data)):
                self.ws.cell(row=self.start_row + 3 + weekends_abbr, column=col_val).style = style3
                self.ws.cell(row=self.start_row + 3 + weekends_abbr, column=col_val).fill = PatternFill(
                    fill_type="solid", fgColor="FFFF99")
                self.ws.cell(row=self.start_row + 3 + weekends_abbr, column=col_val).value = "W"

        # Fill Working days attendance status

    def attendance_data(self):
        att_status = ['P', 'H', 'A', 'L']
        return np.random.choice(att_status, p=[0.7, 0.1, 0.16, 0.04], size=len(self.working_days()))

    def shuffled_data(self):
        emp_list = list(self.attendance_data())
        random.shuffle(emp_list)
        return emp_list

    def fill_attendance(self):
        for att in range(0, len(self.working_days())):
            att_col = self.working_days()[att] + 2
            for att_status in range(0, len(self.read_db())):
                self.ws.cell(row=self.start_row + 3 + att_status, column=att_col).alignment = Alignment(
                    horizontal="center")
                self.ws.cell(row=self.start_row + 3 + att_status, column=att_col).border = Border(
                    left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"),
                    bottom=Side(style="thin"))
                self.ws.cell(row=self.start_row + 3 + att_status, column=att_col).value = self.shuffled_data()[att]

    def save_excel(self):
        # Save the excel file
        self.wb.save(self.filename)
        print("Excel file saved in {}".format(self.filename))


if __name__ == "__main__":
    class_obj = GenerateAttendance()
    class_obj.create_excel_file()
    class_obj.fill_attendance()
    class_obj.save_excel()
